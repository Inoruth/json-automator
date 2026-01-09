from fastapi import FastAPI, UploadFile, File, HTTPException
from fastapi.responses import JSONResponse, HTMLResponse, RedirectResponse, FileResponse
from typing import List, Dict, Any, Tuple
import openpyxl
from io import BytesIO
import json
import os

app = FastAPI()

# ---------- STATS ----------

STATS_FILE = "stats.json"


def load_stats() -> Dict[str, int]:
    """
    Load usage statistics from the JSON file.

    Returns a dictionary with:
    - total: number of conversions (rows + config)
    - rows: number of 'rows' mode conversions
    - config: number of 'config' mode conversions
    """
    if not os.path.exists(STATS_FILE):
        return {"total": 0, "rows": 0, "config": 0}
    with open(STATS_FILE, "r") as f:
        return json.load(f)


def save_stats(stats: Dict[str, int]) -> None:
    """
    Persist usage statistics to the JSON file.

    Parameters
    ----------
    stats : dict
        Dictionary returned by load_stats(), updated with new counts.
    """
    with open(STATS_FILE, "w") as f:
        json.dump(stats, f)


# ---------- UI ROUTES ----------


@app.get("/", response_class=HTMLResponse)
def root():
    """
    Redirect root URL to the main web UI (/app).
    """
    return RedirectResponse(url="/app")


@app.get("/status")
def status():
    """
    Healthcheck endpoint to verify that the app is running.
    """
    return {"status": "ok", "message": "JSON Automator is running 🚀"}


@app.get("/app", response_class=HTMLResponse)
def app_ui():
    """
    Serve the main HTML UI.

    The UI lets the user:
    - upload an Excel file (.xlsx)
    - choose conversion mode (rows / config)
    - see the JSON result
    - download the JSON
    - download an example Excel file
    """
    html_content = """
    <!DOCTYPE html>
    <html lang="en">
    <head>
        <meta charset="UTF-8" />
        <title>JSON Automator – Excel to JSON</title>

        <style>
            body { font-family: system-ui,-apple-system,BlinkMacSystemFont,"Segoe UI",sans-serif;
                   max-width: 960px; margin: 40px auto; padding: 0 16px; color:#111827; }
            h1 { font-size: 2rem; margin-bottom: .5rem; }
            h2 { font-size: 1.3rem; margin-bottom: .5rem; }
            p  { margin: 0.25rem 0 0.75rem 0; }
            .hero { margin-bottom: 16px; }
            .badge { display:inline-block; font-size:.75rem; padding:2px 8px;
                     border-radius:999px; background:#ecfdf5; color:#047857; margin-bottom:8px; }
            .grid { display:grid; grid-template-columns: 1.2fr 1fr; gap:24px; align-items:flex-start; }
            .card { border: 1px solid #e5e7eb; border-radius: 10px; padding: 16px; margin-bottom: 24px; background:white; }
            .row { margin-bottom: 12px; }
            label { display:block; margin-bottom:4px; font-weight:500; }
            .radio-group { display:flex; gap:16px; margin-top:4px; }
            button { padding:8px 16px; border-radius:6px; border:none; cursor:pointer; font-weight:500; }
            #convertBtn { background:#0f766e; color:white; }
            #convertBtn:disabled { background:#9ca3af; cursor:not-allowed; }
            #downloadBtn { background:#2563eb; color:white; margin-left:8px; }
            #downloadBtn:disabled { background:#9ca3af; }
            #exampleBtn { background:#6b7280; color:white; margin-left:8px; }
            pre { background:#111827; color:#e5e7eb; padding:12px; border-radius:8px; overflow-x:auto; font-size:.9rem; }
            .messages { margin-top:8px; color:#92400e; font-size:.9rem; }
            .muted { color:#6b7280; font-size:.85rem; }
            table { border-collapse: collapse; width:100%; font-size:.85rem; }
            th, td { border:1px solid #e5e7eb; padding:6px 8px; }
            th { background:#f9fafb; }
            .notice { background:#eef2ff; border-color:#c7d2fe; }
            .notice strong { font-weight:600; }
        </style>
    </head>

    <body>
        <div class="hero">
            <span class="badge">Free beta</span>
            <h1>JSON Automator</h1>
            <p>Convert Excel config sheets into clean JSON, with basic validation.</p>
        </div>

        <div class="card notice">
            <p>
                <strong>Note:</strong> JSON Automator works best for
                <strong>flat or lightly-nested configuration files</strong>
                (API settings, feature flags, app configs, etc.).
                Support for deep / highly-nested JSON structures is planned.
            </p>
        </div>

        <div class="grid">
            <div>
                <div class="card">
                    <h2>1. Use the tool</h2>

                    <div class="row">
                        <label for="fileInput">Excel file (.xlsx)</label>
                        <input type="file" id="fileInput" accept=".xlsx" />
                    </div>

                    <div class="row">
                        <label>Conversion mode</label>
                        <div class="radio-group">
                            <label><input type="radio" name="mode" value="rows" checked /> Rows (debug)</label>
                            <label><input type="radio" name="mode" value="config" /> Config key/value</label>
                        </div>
                    </div>

                    <div class="row">
                        <button id="convertBtn">Convert</button>
                        <button id="downloadBtn" disabled>Download JSON</button>
                        <button id="exampleBtn" type="button">Download example (.xlsx)</button>
                    </div>
                </div>

                <div class="card">
                    <h2>2. JSON result</h2>
                    <pre id="result">{}</pre>
                    <div id="messages" class="messages"></div>
                </div>
            </div>

            <div>
                <div class="card">
                    <h2>Expected Excel format</h2>
                    <p class="muted">Minimal example:</p>

                    <table>
                        <tr><th>key</th><th>value</th><th>required</th><th>type</th></tr>
                        <tr><td>api_url</td><td>https://api.example.com</td><td>yes</td><td>url</td></tr>
                        <tr><td>timeout</td><td>30</td><td>no</td><td>int</td></tr>
                        <tr><td>use_cache</td><td>true</td><td>no</td><td>bool</td></tr>
                    </table>

                    <p class="muted">Automatic validation (ints, booleans, URLs, missing required fields…).</p>
                </div>
            </div>
        </div>

        <script>
            const btn = document.getElementById('convertBtn');
            const downloadBtn = document.getElementById('downloadBtn');
            const exampleBtn = document.getElementById('exampleBtn');
            const fileInput = document.getElementById('fileInput');
            const resultPre = document.getElementById('result');
            const messagesDiv = document.getElementById('messages');

            let lastResult = null;

            btn.addEventListener('click', async () => {
                const file = fileInput.files[0];
                if (!file) { alert("Please select a .xlsx file first."); return; }

                const mode = document.querySelector('input[name=\\"mode\\"]:checked').value;
                const endpoint = mode === 'rows' ? '/convert' : '/convert/config';

                const formData = new FormData();
                formData.append('file', file);

                btn.disabled = true;
                downloadBtn.disabled = true;
                btn.textContent = "Converting...";
                messagesDiv.textContent = "";
                resultPre.textContent = "{}";

                try {
                    const res = await fetch(endpoint, { method: 'POST', body: formData });
                    const type = res.headers.get('content-type') || '';
                    const data = await res.json();

                    if (!res.ok) {
                        if (type.includes('application/json')) {
                            if (data.detail && data.detail.messages)
                                messagesDiv.textContent = data.detail.messages.join(' | ');
                            else
                                messagesDiv.textContent = JSON.stringify(data.detail || data);
                        } else {
                            messagesDiv.textContent = "Server error while converting.";
                        }
                        return;
                    }

                    lastResult = data;
                    downloadBtn.disabled = false;
                    resultPre.textContent = JSON.stringify(data, null, 2);
                    if (data.messages?.length) {
                        messagesDiv.textContent = data.messages.join(' | ');
                    }
                }
                catch (e) {
                    messagesDiv.textContent = "Network error while calling the API.";
                }
                finally {
                    btn.disabled = false;
                    btn.textContent = "Convert";
                }
            });

            downloadBtn.addEventListener('click', () => {
                if (!lastResult) return;

                const blob = new Blob(
                    [JSON.stringify(lastResult, null, 2)],
                    { type: "application/json" }
                );

                const url = URL.createObjectURL(blob);
                const a = document.createElement("a");
                a.href = url;
                a.download = "result.json";
                a.click();
                URL.revokeObjectURL(url);
            });

            exampleBtn.addEventListener('click', () => {
                window.location.href = "/example";
            });
        </script>
    </body>
    </html>
    """
    return HTMLResponse(content=html_content)


# ---------- EXCEL LOGIC ----------


def parse_excel_to_json(file_bytes: bytes) -> List[Dict[str, Any]]:
    """
    Parse the first sheet of an Excel workbook into a list of row dictionaries.

    Parameters
    ----------
    file_bytes : bytes
        Raw content of the uploaded .xlsx file.

    Returns
    -------
    list[dict]
        One dict per non-empty row. Keys come from the header row (row 1),
        values from subsequent rows.

    Raises
    ------
    HTTPException (400)
        If the file is not a valid Excel workbook or the header row is empty.
    """
    try:
        workbook = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="Could not read Excel file.")

    sheet = workbook.active
    headers = [str(c.value) if c.value is not None else "" for c in sheet[1]]

    if not any(headers):
        raise HTTPException(status_code=400, detail="First row must contain headers.")

    rows: List[Dict[str, Any]] = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        d: Dict[str, Any] = {}
        for idx, v in enumerate(row):
            col_name = headers[idx] if idx < len(headers) else f"col_{idx}"

            # special handling: if this is the 'required' column and the cell is empty,
            # normalize it to "no" instead of leaving it as null
            if col_name == "required" and (v is None or str(v).strip() == ""):
                d[col_name] = "no"
            else:
                d[col_name] = v

        # Ignore fully empty rows
        if any(v not in (None, "") for v in d.values()):
            rows.append(d)

    return rows


def rows_to_config_key_value(rows: List[Dict[str, Any]]) -> Tuple[Dict[str, Any], List[str]]:
    """
    Convert a list of row dictionaries into a config object.

    Expected columns (case-sensitive):
    - key (required)
    - value (required)
    - required (optional, yes/no)
    - type (optional, string/int/bool/url)

    The function:
    - builds a config dict: {key: converted_value}
    - validates required fields and basic types
    - collects warnings / errors in a list of messages

    Parameters
    ----------
    rows : list[dict]
        Output of parse_excel_to_json().

    Returns
    -------
    config : dict
        Final configuration object.
    messages : list[str]
        Validation messages (warnings / errors). Can be empty.
    """
    config: Dict[str, Any] = {}
    messages: List[str] = []

    if not rows:
        messages.append("No data rows found.")
        return config, messages

    first = rows[0]
    if "key" not in first or "value" not in first:
        messages.append("Columns 'key' and 'value' are required.")
        return config, messages

    seen = set()

    for idx, row in enumerate(rows, start=2):
        key_raw = row.get("key")
        value = row.get("value")

        # If 'required' cell is empty or normalized to "no", treat accordingly
        required = (str(row.get("required", "")).strip().lower() or "no")
        value_type = str(row.get("type", "string")).strip().lower()

        if not key_raw or str(key_raw).strip() == "":
            messages.append(f"Row {idx}: empty key — ignored.")
            continue

        key = str(key_raw).strip()

        if key in seen:
            messages.append(f"Row {idx}: duplicate key '{key}' — overwriting previous value.")
        seen.add(key)

        if required == "yes" and (value is None or str(value) == ""):
            messages.append(f"Row {idx}: missing required value for '{key}'.")

        converted = value
        if value_type == "int":
            try:
                converted = int(value)
            except Exception:
                messages.append(f"Row {idx}: '{key}' should be an integer.")
        elif value_type == "bool":
            s = str(value).lower()
            if s in ["true", "1", "yes"]:
                converted = True
            elif s in ["false", "0", "no"]:
                converted = False
            else:
                messages.append(f"Row {idx}: '{key}' should be a boolean (true/false, yes/no).")
        elif value_type == "url":
            if not str(value).startswith(("http://", "https://")):
                messages.append(f"Row {idx}: '{key}' should be a valid URL (http/https).")

        config[key] = converted

    if not config:
        messages.append("No valid entries were generated from the sheet.")

    return config, messages


# ---------- API ENDPOINTS ----------


@app.post("/convert")
async def convert_excel_to_json_api(file: UploadFile = File(...)):
    """
    API endpoint for the 'rows' mode.

    - Accepts an Excel file (.xlsx)
    - Parses it into a list of row dictionaries
    - Updates usage statistics
    - Returns: {"rows": [...]}
    """
    if not file.filename.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Only .xlsx files are supported.")

    rows = parse_excel_to_json(await file.read())

<<<<<<< Updated upstream
    stats = load_stats()
    stats["total"] += 1
    stats["rows"] += 1
    save_stats(stats)

    return JSONResponse(content={"rows": rows})
=======
    # Initialize result container
    result = {"sheets": {}, "messages": []}

    # Process each sheet in the workbook
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        headers = [cell.value for cell in sheet[1]]  # Get header row

        # Convert rows to JSON objects
        rows = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            obj = {headers[i]: row[i] for i in range(len(headers))}
            rows.append(obj)

        # Add to result
        result["sheets"][sheet_name] = rows

    # Handle different conversion modes
    if mode == "rows":
        # "Rows" mode just returns the raw rows as JSON
        result = {"data": result["sheets"], "messages": result["messages"]}
    elif mode == "config":
        # "Config" mode expects two columns: "key" and "value"
        for sheet_name, rows in result["sheets"].items():
            for row in rows:
                if "key" in row and "value" in row:
                    # Simple key/value pair
                    result["messages"].append(f"Config: {row['key']} = {row['value']}")
                else:
                    result["messages"].append(f"Row in sheet '{sheet_name}' is missing 'key' or 'value': {row}")
        result = {"data": {}, "messages": result["messages"]}
    elif mode == "config_schema":
        # "Schema-based config" mode
        if not schema_file:
            raise HTTPException(status_code=400, detail="Schema file is required for 'Schema-based config' mode.")

        # Load schema JSON
        try:
            schema_bytes = await schema_file.read();
            schema = json.loads(schema_bytes.decode("utf-8"));
        except Exception:
            raise HTTPException(status_code=400, detail="Schema file is not valid JSON. Upload a .json with the expected structure.");

        # Basic schema checks
        if not isinstance(schema, dict):
            raise HTTPException(status_code=400, detail="Schema root must be a JSON object.");
        if "columns" not in schema or not isinstance(schema.get("columns"), dict):
            raise HTTPException(status_code=400, detail="Schema must define 'columns' as an object with header aliases.");
        if "keys" not in schema or not isinstance(schema.get("keys"), dict):
            raise HTTPException(status_code=400, detail="Schema must define 'keys' as an object with rules per key.");

        # Build alias map for headers
        alias_map = {};
        for canonical, aliases in schema["columns"].items():
            if isinstance(aliases, list):
                for a in aliases:
                    alias_map[str(a)] = canonical;
            alias_map[canonical] = canonical;  # map canonical to itself

        # Normalize headers for each sheet’s rows
        normalized_sheets = {};
        for sheet_name, rows in result["sheets"].items():
            if not rows:
                normalized_sheets[sheet_name] = [];
                continue
            # Use headers from first row keys
            original_headers = list(rows[0].keys());
            header_map = {h: alias_map.get(h, h) for h in original_headers};
            norm_rows = [];
            for r in rows:
                nr = {};
                for k, v in r.items():
                    nr[header_map.get(k, k)] = v;
                norm_rows.append(nr);
            normalized_sheets[sheet_name] = norm_rows;

        # Prepare key rules and alias resolution
        key_rules = schema.get("keys", {});
        alias_to_key = {};
        for k, rule in key_rules.items():
            if isinstance(rule, dict):
                for a in (rule.get("aliases", []) or []):
                    alias_to_key[str(a)] = k;

        allow_extra = bool(schema.get("allow_extra_keys", False));

        # Helper: type conversion
        def convert_type(val: Any, expected: str, row_idx: int, key: str) -> Any:
            t = (expected or "string").lower();
            if t == "int":
                try:
                    return int(val);
                except Exception:
                    result["messages"].append(f"Row {row_idx}: '{key}' expects an integer. Example: 0, 10, 300.");
                    return val;
            if t == "bool":
                s = str(val).strip().lower();
                if s in ["true", "1", "yes"]:
                    return True;
                if s in ["false", "0", "no"]:
                    return False;
                result["messages"].append(f"Row {row_idx}: '{key}' expects a boolean. Use true/false or yes/no.");
                return val;
            if t == "url":
                if str(val).startswith(("http://", "https://")):
                    return val;
                result["messages"].append(f"Row {row_idx}: '{key}' expects a URL starting with http:// or https://.");
                return val;
            return val;  # string/default

        # Helper: set nested via dot-notation
        def set_nested(dct: Dict[str, Any], dotted: str, val: Any) -> None:
            parts = dotted.split(".");
            cur = dct;
            for p in parts[:-1]:
                if p not in cur or not isinstance(cur[p], dict):
                    cur[p] = {};
                cur = cur[p];
            cur[parts[-1]] = val;

        # Build config from normalized rows (first sheet wins; keep current structure minimal)
        final_config: Dict[str, Any] = {};
        seen_keys = set();

        # Use all sheets, row numbers restart per sheet (keeps current behavior simple)
        for sheet_name, rows in normalized_sheets.items():
            for idx, row in enumerate(rows, start=2):
                raw_key = row.get("key");
                value = row.get("value");

                if not raw_key or str(raw_key).strip() == "":
                    result["messages"].append(f"Row {idx}: The 'key' cell is empty — this row was ignored. Fill in a key name.");
                    continue;

                key_name = str(raw_key).strip();
                canonical_key = alias_to_key.get(key_name, key_name);
                rule = key_rules.get(canonical_key, {});

                # Required handling (Excel 'required' or schema 'required')
                req_excel = str(row.get("required", "")).strip().lower() == "yes";
                req_schema = bool(rule.get("required", False));
                is_required = req_excel or req_schema;

                # Default handling (schema default if Excel value empty)
                if (value is None or str(value).strip() == "") and ("default" in rule):
                    value = rule["default"];

                # Required check after default
                if is_required and (value is None or str(value).strip() == ""):
                    result["messages"].append(f"Row {idx}: Missing required value for '{canonical_key}'. Enter a value in 'value' or provide a default in the schema.");

                # Type validation (schema type preferred, fallback to Excel 'type')
                expected_type = str(rule.get("type", row.get("type", "string"))).strip().lower();
                converted = convert_type(value, expected_type, idx, canonical_key);

                # Duplicate keys warning
                if canonical_key in seen_keys:
                    result["messages"].append(f"Row {idx}: Duplicate key '{canonical_key}' — the later value overwrote the earlier one. Use unique keys.");
                seen_keys.add(canonical_key);

                # Extra key handling
                if canonical_key not in key_rules and not allow_extra:
                    result["messages"].append(f"Row {idx}: Key '{canonical_key}' is not defined in the schema. It was ignored because 'allow_extra_keys' is false.");
                    continue;  # do not include this key in the final config when extras are disallowed

                # Dot-notation nesting
                set_nested(final_config, canonical_key, converted);

        # Missing required keys from schema
        for skey, rule in key_rules.items():
            if rule.get("required", False) and skey not in seen_keys:
                result["messages"].append(f"Schema required key missing: '{skey}'. Add a row with this key or provide a default in the schema.");

        # Return in the same shape as other modes
        if not final_config:
            return JSONResponse(status_code=400, content={"data": {}, "messages": result["messages"]});
        result = {"data": final_config, "messages": result["messages"]};
    else:
        raise HTTPException(status_code=400, detail="Invalid mode. Choose 'rows', 'config', or 'config_schema'.");

    return JSONResponse(content=result);


@app.get("/example")
def get_example_file():
    """
    Serve an example Excel file for download.
    """
    file_path = "example_config.xlsx";
    if not os.path.exists(file_path):
        raise HTTPException(status_code=404, detail="Example file not found.");
    return FileResponse(file_path, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", filename="example_config.xlsx");


@app.get("/stats")
def get_stats():
    """
    Get usage statistics (total conversions, rows mode, config mode).
    """
    stats = load_stats();
    return {"total_conversions": stats["total"], "rows_mode": stats["rows"], "config_mode": stats["config"]};


@app.on_event("startup")
def startup_event():
    """
    Application startup event to initialize any required resources.
    """
    # Ensure the stats file exists
    if not os.path.exists(STATS_FILE):
        save_stats({"total": 0, "rows": 0, "config": 0});
>>>>>>> Stashed changes


@app.post("/convert/config")
async def convert_excel_to_config_api(file: UploadFile = File(...)):
    """
    API endpoint for the 'config' mode.

    - Accepts an Excel file (.xlsx)
    - Parses it into rows
    - Builds a config object using rows_to_config_key_value()
    - Updates usage statistics
    - On error: returns HTTP 400 with validation messages
    """
<<<<<<< Updated upstream
    if not file.filename.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Only .xlsx files are supported.")

    rows = parse_excel_to_json(await file.read())
    config, messages = rows_to_config_key_value(rows)

    stats = load_stats()
    stats["total"] += 1
    stats["config"] += 1
    save_stats(stats)

    if not config:
        # All messages are returned to the client when nothing valid is produced
        raise HTTPException(status_code=400, detail={"messages": messages})

    return JSONResponse(content={"config": config, "messages": messages})
=======
    return await convert(file=file, mode="config");
>>>>>>> Stashed changes


# ---------- EXAMPLE EXCEL ----------


@app.get("/example")
def download_example():
    """
    Download an example Excel file that matches the expected format.

    The file 'example.xlsx' must be present at the project root
    (same directory as main.py).
    """
<<<<<<< Updated upstream
    example_path = "example.xlsx"
    if not os.path.exists(example_path):
        raise HTTPException(status_code=404, detail="Example file not found.")
    return FileResponse(example_path, filename="example.xlsx")


# ---------- ADMIN ----------


@app.get("/_admin/stats")
def get_stats():
    """
    Return global usage statistics.

    This endpoint is open while the app is in beta, but it can be
    restricted later if needed.
    """
    return load_stats()
=======
    return await convert(file=file, mode="config_schema", schema_file=schema_file);


@app.get("/_admin/stats")
@app.get("/_admin/stats/")
def admin_stats_alias():
    """
    Backward-compatible stats endpoint (legacy path).
    Returns the same payload as GET /stats.
    """
    return get_stats()
>>>>>>> Stashed changes
